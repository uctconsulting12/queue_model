"""
Excel Logger for Queue Monitoring System
Saves inference results to timestamp.xlsx, one row per frame.
Supports batch writing for better performance.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "timestamp.xlsx"

HEADERS = [
    "Frame_Id",
    "Time_stamp",
    "Queue_Count",
    "Queue_Names",
    "Queue_Lengths",
    "Front_Person_Wait",
    "Avg_Wait_Time",
    "Status",
    "Should_Alert",
    "Total_Customer_Visited",
    "Queue_Customer_Visited",
    "Total_People_Detected",
    "People_IDs",
    "Queue_Assignments",
    "Entry_Times",
    "People_Wait_Times",
    "Processing_Time_ms",
    "Active_Tracks",
    "Unique_Persons",
    "Total_Tracks_Created",
]

COLUMN_WIDTHS = [
    18, 22, 14, 30, 20, 22, 20, 35, 16, 22,
    25, 22, 30, 25, 25, 30, 20, 16, 16, 20,
]

HEADER_FILL   = PatternFill("solid", start_color="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT     = Font(name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="EBF0FA")
ALERT_FILL    = PatternFill("solid", start_color="FFD7D7")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


class ExcelLogger:
    """Excel logger that supports batch writing"""

    def __init__(self, excel_path=EXCEL_PATH):
        self.excel_path = excel_path
        self.results_df = pd.DataFrame(columns=HEADERS)
        self.frame_count = 0

    def _list_to_str(self, value) -> str:
        """Convert list to comma-separated string"""
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value) if value is not None else ""

    def add_result(self, result: dict):
        """
        Add one inference result to DataFrame (in-memory)
        """
        tracker = result.get("Tracker_Stats", {})

        row_data = {
            "Frame_Id": result.get("Frame_Id", self.frame_count + 1),
            "Time_stamp": result.get("Time_stamp", ""),
            "Queue_Count": result.get("Queue_Count", 0),
            "Queue_Names": self._list_to_str(result.get("Queue_Name", [])),
            "Queue_Lengths": self._list_to_str(result.get("Queue_Length", [])),
            "Front_Person_Wait": self._list_to_str(result.get("Front_person_Wt", [])),
            "Avg_Wait_Time": self._list_to_str(result.get("Average_wt_time", [])),
            "Status": self._list_to_str(result.get("Status", [])),
            "Should_Alert": self._list_to_str(result.get("Should_Alert", [])),
            "Total_Customer_Visited": result.get("Total_Customer_Visited", 0),
            "Queue_Customer_Visited": self._list_to_str(result.get("Queue_Customer_Visited", [])),
            "Total_People_Detected": result.get("Total_people_detected", 0),
            "People_IDs": self._list_to_str(result.get("People_ids", [])),
            "Queue_Assignments": self._list_to_str(result.get("Queue_Assignment", [])),
            "Entry_Times": self._list_to_str(result.get("Entry_time", [])),
            "People_Wait_Times": self._list_to_str(result.get("People_wt_time", [])),
            "Processing_Time_ms": result.get("Processing_Time_ms", 0),
            "Active_Tracks": tracker.get("active_tracks", 0),
            "Unique_Persons": tracker.get("unique_persons", 0),
            "Total_Tracks_Created": tracker.get("total_tracks_created", 0),
        }

        # Use pandas concat instead of append (append is deprecated)
        self.results_df = pd.concat([self.results_df, pd.DataFrame([row_data])], ignore_index=True)
        self.frame_count += 1

    def save_to_excel(self):
        """
        Write all accumulated results to Excel file with formatting
        """
        if self.results_df.empty:
            print("No results to save")
            return

        # Create workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Queue Monitoring Log"

        # Apply header formatting
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        for col_idx, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, (_, row_data) in enumerate(self.results_df.iterrows(), start=2):
            has_alert = "True" in str(row_data["Should_Alert"]) or "1" in str(row_data["Should_Alert"])
            alt_row = (row_idx % 2 == 0)

            for col_idx, header in enumerate(HEADERS, start=1):
                value = row_data[header]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER

                # Center align specific columns
                if col_idx in (1, 2, 3, 10, 12, 17, 18, 19, 20):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT

                # Apply row coloring
                if has_alert:
                    cell.fill = ALERT_FILL
                elif alt_row:
                    cell.fill = ALT_FILL

        # Save workbook
        wb.save(self.excel_path)
        print(f"\n✓ Results saved to: {self.excel_path}")
        print(f"  Total rows written: {len(self.results_df)}")

    def clear(self):
        """Clear the DataFrame for new batch"""
        self.results_df = pd.DataFrame(columns=HEADERS)
        self.frame_count = 0


# For backward compatibility, create a global instance
_default_logger = ExcelLogger()

def log_result(result: dict):
    """
    Legacy function for backward compatibility
    """
    _default_logger.add_result(result)

def save_results():
    """
    Save all accumulated results to Excel
    """
    _default_logger.save_to_excel()

def clear_results():
    """
    Clear accumulated results
    """
    _default_logger.clear()